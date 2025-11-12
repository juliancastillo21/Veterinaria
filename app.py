from flask import Flask, render_template, request, redirect, url_for, jsonify
from openpyxl import Workbook, load_workbook
from datetime import datetime
import base64
from io import BytesIO
from PIL import Image, ImageOps

app = Flask(__name__)

# Configuración
EXCEL_FILE = 'registros_vacas.xlsx'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máximo

def allowed_file(filename):
    """Verifica si el archivo tiene una extensión permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def procesar_imagen_a_base64(file_storage, max_base64_len: int = 32000) -> str:
    # Tamaños y calidades a intentar
    tamanos = [800, 720, 640, 560, 480]
    calidades = [85, 80, 75, 70, 65, 60, 55, 50]

    # Asegurarnos de leer desde el inicio del stream
    try:
        file_storage.stream.seek(0)
    except Exception:
        pass

    # Abrir imagen con PIL
    img = Image.open(file_storage.stream)

    # Corregir orientación y convertir a RGB
    img = ImageOps.exif_transpose(img)
    if img.mode != 'RGB':
        img = img.convert('RGB')

    for max_px in tamanos:
        # Redimensionar manteniendo proporción
        copia = img.copy()
        copia.thumbnail((max_px, max_px), Image.LANCZOS)

        for q in calidades:
            buf = BytesIO()
            copia.save(buf, format='JPEG', quality=q, optimize=True)
            raw = buf.getvalue()
            b64 = base64.b64encode(raw).decode('utf-8')
            if len(b64) <= max_base64_len:
                return b64

    # Si no se pudo cumplir el límite, devolver la versión más comprimida posible
    # Último intento: 480px, calidad 50
    buf = BytesIO()
    img_small = img.copy()
    img_small.thumbnail((480, 480), Image.LANCZOS)
    img_small.save(buf, format='JPEG', quality=50, optimize=True)
    return base64.b64encode(buf.getvalue()).decode('utf-8')

def guardar_en_excel(datos):
    """Guarda los datos en el archivo Excel"""
    # Cargar el archivo existente
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    ws.append([
        datos['fecha_hora'],
        datos['nombre_ordenador'],
        datos['id_vaca'],
        datos['nombre_vaca'],
        datos['litros'],
        datos['imagen_base64'],
        datos['edad'],
        datos['estado_productivo'],
        datos['vaca_parida'],
        datos['vaca_seca'],
        datos['numero_crias'],
        datos['numero_parto'],
        datos.get('vacunas', ''),
        datos.get('enfermedades', '')
    ])
    
    wb.save(EXCEL_FILE)

def ensure_workbook_and_headers():
    """Asegura que el archivo Excel exista y tenga encabezados mínimos en la hoja principal y la hoja 'Crias'."""
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Registros'
        ws.append([
            'FechaHora','Ordeñador','ID Vaca','Nombre Vaca','Litros','Imagen Base64','Edad','Estado','Parida','Seca','Nº Crías','Nº Parto','Vacunas','Enfermedades'
        ])
    # hoja principal
    ws = wb.active
    if ws.max_row == 1 and ws.max_column < 5:
        ws.append([
            'FechaHora','Ordeñador','ID Vaca','Nombre Vaca','Litros','Imagen Base64','Edad','Estado','Parida','Seca','Nº Crías','Nº Parto','Vacunas','Enfermedades'
        ])
    # hoja crias
    if 'Crias' not in wb.sheetnames:
        ws_c = wb.create_sheet('Crias')
        ws_c.append(['FechaRegistro','MadreID','MadreNombre','CriaID','CriaNombre','FechaNacimiento','Sexo','Observaciones'])
    wb.save(EXCEL_FILE)
    wb.close()

def get_unique_cows():
    """Devuelve lista de vacas únicas (ID, Nombre) desde la hoja principal."""
    ensure_workbook_and_headers()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    seen = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        cow_id = str(row[2])
        name = row[3] or ''
        # Conserva el último nombre visto
        seen[cow_id] = name
    wb.close()
    # Lista ordenada por ID
    return sorted([{ 'id': cid, 'nombre': seen[cid] } for cid in seen.keys()], key=lambda x: x['id'])

def get_crias():
    """Lee la hoja 'Crias' y devuelve todas las crías."""
    ensure_workbook_and_headers()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['Crias']
    crias = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[1]:
            continue
        crias.append({
            'fila': idx,
            'fecha_registro': row[0],
            'madre_id': row[1],
            'madre_nombre': row[2],
            'cria_id': row[3],
            'cria_nombre': row[4],
            'fecha_nacimiento': row[5],
            'sexo': row[6],
            'observaciones': row[7] or ''
        })
    wb.close()
    return crias

def add_cria(madre_id: str, madre_nombre: str, cria_id: str, cria_nombre: str, fecha_nac: str, sexo: str, obs: str):
    ensure_workbook_and_headers()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['Crias']
    ws.append([
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        madre_id,
        madre_nombre,
        cria_id,
        cria_nombre,
        fecha_nac,
        sexo,
        obs
    ])
    wb.save(EXCEL_FILE)
    wb.close()

@app.route('/')
def index():
    """Página de inicio"""
    return render_template('inicio.html')

@app.route('/crias')
def crias_view():
    """Vista para gestionar crías: lista y formulario de alta."""
    try:
        vacas = get_unique_cows()
        crias = get_crias()
        return render_template('crias.html', vacas=vacas, crias=crias)
    except Exception as e:
        return render_template('crias.html', vacas=[], crias=[], error=f"Error: {str(e)}")

@app.route('/crias/guardar', methods=['POST'])
def crias_guardar():
    try:
        madre = request.form.get('madre')  # formato: id|nombre
        cria_id = request.form.get('cria_id')
        cria_nombre = request.form.get('cria_nombre')
        fecha_nac = request.form.get('fecha_nacimiento')
        sexo = request.form.get('sexo')
        obs = request.form.get('observaciones') or ''

        if not madre or not cria_id or not cria_nombre or not fecha_nac or not sexo:
            return redirect(url_for('crias_view') + '?error=Faltan campos requeridos')

        if '|' in madre:
            madre_id, madre_nombre = madre.split('|', 1)
        else:
            # fallback: buscar nombre por ID
            madre_id = madre
            madre_nombre = next((v['nombre'] for v in get_unique_cows() if v['id'] == madre_id), '')

        add_cria(madre_id, madre_nombre, cria_id, cria_nombre, fecha_nac, sexo, obs)
        return redirect(url_for('crias_view') + '?success=true')
    except Exception as e:
        return redirect(url_for('crias_view') + f'?error={str(e)}')

@app.route('/formulario')
def formulario():
    """Página del formulario de registro"""
    return render_template('formulario.html', edicion=False, form_action=url_for('guardar'))

@app.route('/editar/<int:fila>')
def editar(fila: int):
    """Carga un registro para editarlo y reutiliza el formulario."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        row = ws[fila]
        # Convertir a values
        valores = [cell.value for cell in row]
        wb.close()

        # Campos seguros por índice
        vacunas_str = valores[12] if len(valores) > 12 else ''
        enfermedades_str = valores[13] if len(valores) > 13 else ''
        vacunas_sel = [v.strip() for v in (vacunas_str or '').split(',') if v and v.strip()]
        enfermedades_sel = [e.strip() for e in (enfermedades_str or '').split(',') if e and e.strip()]

        registro = {
            'fila': fila,
            'fecha_hora': valores[0],
            'nombre_ordenador': valores[1],
            'id_vaca': valores[2],
            'nombre_vaca': valores[3],
            'litros': valores[4],
            'imagen_base64': valores[5],
            'edad': valores[6],
            'estado_productivo': valores[7],
            'vaca_parida': valores[8],
            'vaca_seca': valores[9],
            'numero_crias': valores[10],
            'numero_parto': valores[11],
            'vacunas': vacunas_str,
            'enfermedades': enfermedades_str
        }

        return render_template(
            'formulario.html',
            edicion=True,
            form_action=url_for('actualizar', fila=fila),
            registro=registro,
            vacunas_sel=vacunas_sel,
            enfermedades_sel=enfermedades_sel
        )
    except Exception as e:
        return redirect(url_for('registros') + f'?error=No se pudo cargar el registro: {str(e)}')

@app.route('/actualizar/<int:fila>', methods=['POST'])
def actualizar(fila: int):
    """Actualiza un registro existente en la fila dada."""
    try:
        # Obtener datos del formulario
        nombre_ordenador = request.form.get('nombre_ordenador')
        id_vaca = request.form.get('id_vaca')
        nombre_vaca = request.form.get('nombre_vaca')
        edad = request.form.get('edad')
        estado_productivo = request.form.get('estado_productivo')
        vaca_parida = request.form.get('vaca_parida')
        vaca_seca = request.form.get('vaca_seca')
        numero_crias = request.form.get('numero_crias')
        numero_parto = request.form.get('numero_parto')
        litros = request.form.get('litros')

        vacunas_list = request.form.getlist('vacunas')
        enfermedades_list = request.form.getlist('enfermedades')
        if 'Ninguna' in enfermedades_list:
            enfermedades_list = ['Ninguna']
        vacunas_str = ', '.join(vacunas_list) if vacunas_list else ''
        enfermedades_str = ', '.join(enfermedades_list) if enfermedades_list else 'Ninguna'

        # Validación básica
        if not all([nombre_ordenador, id_vaca, nombre_vaca, edad, estado_productivo, 
                    vaca_parida, vaca_seca, numero_crias, numero_parto, litros]):
            return redirect(url_for('editar', fila=fila) + '?error=Faltan campos requeridos')

        # Cargar Excel y obtener imagen existente por si no se reemplaza
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        imagen_existente = ws.cell(row=fila, column=6).value  # Columna 6 = imagen_base64

        # Procesar foto si se envió una nueva
        imagen_base64 = imagen_existente
        if 'foto' in request.files:
            foto = request.files['foto']
            if foto and foto.filename and allowed_file(foto.filename):
                imagen_base64 = procesar_imagen_a_base64(foto)

        # Escribir valores en la fila (respetando el orden de columnas actual)
        ws.cell(row=fila, column=1, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=fila, column=2, value=nombre_ordenador)
        ws.cell(row=fila, column=3, value=id_vaca)
        ws.cell(row=fila, column=4, value=nombre_vaca)
        ws.cell(row=fila, column=5, value=float(litros))
        ws.cell(row=fila, column=6, value=imagen_base64)
        ws.cell(row=fila, column=7, value=int(edad))
        ws.cell(row=fila, column=8, value=estado_productivo)
        ws.cell(row=fila, column=9, value=vaca_parida)
        ws.cell(row=fila, column=10, value=vaca_seca)
        ws.cell(row=fila, column=11, value=int(numero_crias))
        ws.cell(row=fila, column=12, value=int(numero_parto))
        ws.cell(row=fila, column=13, value=vacunas_str)
        ws.cell(row=fila, column=14, value=enfermedades_str)

        wb.save(EXCEL_FILE)
        wb.close()

        return redirect(url_for('registros') + '?success=edit')
    except Exception as e:
        return redirect(url_for('editar', fila=fila) + f'?error={str(e)}')

@app.route('/registros')
def registros():
    """Página para ver los registros"""
    try:
        # Leer el archivo Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Obtener todos los registros (excepto la fila de encabezados)
        registros_data = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0]:  # Si hay fecha y hora (primera columna no está vacía)
                row_len = len(row)
                vacunas = row[12] if row_len > 12 else ''
                enfermedades = row[13] if row_len > 13 else ''
                registros_data.append({
                    'fila': idx,
                    'fecha_hora': row[0],
                    'nombre_ordenador': row[1],
                    'id_vaca': row[2],
                    'nombre_vaca': row[3],
                    'litros': row[4],
                    'imagen_base64': row[5],
                    'edad': row[6],
                    'estado_productivo': row[7],
                    'vaca_parida': row[8],
                    'vaca_seca': row[9],
                    'numero_crias': row[10],
                    'numero_parto': row[11],
                    'vacunas': vacunas,
                    'enfermedades': enfermedades
                })
        
        wb.close()
        return render_template('registros.html', registros=registros_data)
    except FileNotFoundError:
        return render_template('registros.html', registros=[], error="No se encontró el archivo de registros")
    except Exception as e:
        return render_template('registros.html', registros=[], error=f"Error al leer registros: {str(e)}")

@app.route('/api/registro/<int:fila>')
def api_registro(fila: int):
    """Devuelve un registro en formato JSON para edición modal."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        row = ws[fila]
        valores = [cell.value for cell in row]
        wb.close()

        row_len = len(valores)
        vacunas = valores[12] if row_len > 12 else ''
        enfermedades = valores[13] if row_len > 13 else ''

        data = {
            'fila': fila,
            'fecha_hora': valores[0],
            'nombre_ordenador': valores[1],
            'id_vaca': valores[2],
            'nombre_vaca': valores[3],
            'litros': valores[4],
            'imagen_base64': valores[5],
            'edad': valores[6],
            'estado_productivo': valores[7],
            'vaca_parida': valores[8],
            'vaca_seca': valores[9],
            'numero_crias': valores[10],
            'numero_parto': valores[11],
            'vacunas': vacunas or '',
            'enfermedades': enfermedades or ''
        }
        return jsonify(success=True, registro=data)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 400

@app.route('/estadisticas')
def estadisticas():
    """Página de estadísticas"""
    try:
        # Leer el archivo Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Obtener todos los registros
        registros = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Si hay fecha
                registros.append({
                    'fecha_hora': row[0],
                    'nombre_ordenador': row[1],
                    'id_vaca': row[2],
                    'nombre_vaca': row[3],
                    'litros': float(row[4]) if row[4] else 0,
                    'edad': int(row[6]) if row[6] else 0,
                    'estado_productivo': row[7],
                    'vaca_parida': row[8],
                    'vaca_seca': row[9],
                    'numero_crias': int(row[10]) if row[10] else 0,
                    'numero_parto': int(row[11]) if row[11] else 0
                })
        
        wb.close()
        
        # Calcular estadísticas
        total_vacas = len(registros)
        
        if total_vacas > 0:
            # Producción de leche
            total_litros = sum(r['litros'] for r in registros)
            promedio_litros = total_litros / total_vacas
            max_litros = max(r['litros'] for r in registros)
            min_litros = min(r['litros'] for r in registros)
            
            # Estado productivo
            productivas = sum(1 for r in registros if r['estado_productivo'] == 'Productiva')
            no_productivas = sum(1 for r in registros if r['estado_productivo'] == 'No Productiva')
            en_reposo = sum(1 for r in registros if r['estado_productivo'] == 'En Reposo')
            
            # Edad
            promedio_edad = sum(r['edad'] for r in registros) / total_vacas
            
            # Vacas paridas y secas
            vacas_paridas = sum(1 for r in registros if r['vaca_parida'] == 'Sí')
            vacas_secas = sum(1 for r in registros if r['vaca_seca'] == 'Sí')
            
            # Reproducción
            total_crias = sum(r['numero_crias'] for r in registros)
            promedio_crias = total_crias / total_vacas
            promedio_partos = sum(r['numero_parto'] for r in registros) / total_vacas
            
            # Top 5 mejores productoras
            top_productoras = sorted(registros, key=lambda x: x['litros'], reverse=True)[:5]
            
            # Producción por ordeñador
            ordenadores = {}
            for r in registros:
                ord_nombre = r['nombre_ordenador']
                if ord_nombre in ordenadores:
                    ordenadores[ord_nombre]['total'] += r['litros']
                    ordenadores[ord_nombre]['count'] += 1
                else:
                    ordenadores[ord_nombre] = {'total': r['litros'], 'count': 1}
            
            stats = {
                'total_vacas': total_vacas,
                'total_litros': round(total_litros, 2),
                'promedio_litros': round(promedio_litros, 2),
                'max_litros': round(max_litros, 2),
                'min_litros': round(min_litros, 2),
                'productivas': productivas,
                'no_productivas': no_productivas,
                'en_reposo': en_reposo,
                'promedio_edad': round(promedio_edad, 1),
                'vacas_paridas': vacas_paridas,
                'vacas_secas': vacas_secas,
                'total_crias': total_crias,
                'promedio_crias': round(promedio_crias, 1),
                'promedio_partos': round(promedio_partos, 1),
                'top_productoras': top_productoras,
                'ordenadores': ordenadores
            }
        else:
            stats = None
        
        return render_template('estadisticas.html', stats=stats)
    except FileNotFoundError:
        return render_template('estadisticas.html', stats=None, error="No se encontró el archivo de registros")
    except Exception as e:
        return render_template('estadisticas.html', stats=None, error=f"Error al calcular estadísticas: {str(e)}")

@app.route('/guardar', methods=['POST'])
def guardar():
    """Procesa el formulario y guarda los datos"""
    try:
        # Obtener datos del formulario
        nombre_ordenador = request.form.get('nombre_ordenador')
        id_vaca = request.form.get('id_vaca')
        nombre_vaca = request.form.get('nombre_vaca')
        edad = request.form.get('edad')
        estado_productivo = request.form.get('estado_productivo')
        vaca_parida = request.form.get('vaca_parida')
        vaca_seca = request.form.get('vaca_seca')
        numero_crias = request.form.get('numero_crias')
        numero_parto = request.form.get('numero_parto')
        litros = request.form.get('litros')
        vacunas_list = request.form.getlist('vacunas')
        enfermedades_list = request.form.getlist('enfermedades')
        # Normalizar enfermedades: 'Ninguna' es excluyente
        if 'Ninguna' in enfermedades_list:
            enfermedades_list = ['Ninguna']
        vacunas_str = ', '.join(vacunas_list) if vacunas_list else ''
        enfermedades_str = ', '.join(enfermedades_list) if enfermedades_list else 'Ninguna'
        
        # Verificar que todos los campos estén presentes
        if not all([nombre_ordenador, id_vaca, nombre_vaca, edad, estado_productivo, 
                    vaca_parida, vaca_seca, numero_crias, numero_parto, litros]):
            return redirect(url_for('formulario') + '?error=Faltan campos requeridos')
        
        # Procesar la foto
        if 'foto' not in request.files:
            return redirect(url_for('formulario') + '?error=No se encontró la foto')
        
        foto = request.files['foto']
        
        if foto.filename == '':
            return redirect(url_for('formulario') + '?error=No se seleccionó ninguna foto')
        
        if foto and allowed_file(foto.filename):
            # Procesar y comprimir la imagen en base64 segura para Excel
            imagen_base64 = procesar_imagen_a_base64(foto)
            # Log simple en consola para depuración de longitud
            print(f"Longitud base64 generada: {len(imagen_base64)} caracteres")
            
            # Preparar datos para guardar
            datos = {
                'fecha_hora': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'nombre_ordenador': nombre_ordenador,
                'id_vaca': id_vaca,
                'nombre_vaca': nombre_vaca,
                'edad': int(edad),
                'estado_productivo': estado_productivo,
                'vaca_parida': vaca_parida,
                'vaca_seca': vaca_seca,
                'numero_crias': int(numero_crias),
                'numero_parto': int(numero_parto),
                'litros': float(litros),
                'imagen_base64': imagen_base64,
                'vacunas': vacunas_str,
                'enfermedades': enfermedades_str
            }
            
            # Guardar en Excel
            guardar_en_excel(datos)
            
            return redirect(url_for('formulario') + '?success=true')
        else:
            return redirect(url_for('formulario') + '?error=Formato de imagen no permitido')
            
    except Exception as e:
        return redirect(url_for('formulario') + f'?error={str(e)}')

if __name__ == '__main__':
    print("=" * 50)
    print("🐄 Servidor de Veterinaria iniciado")
    print("=" * 50)
    print(f"📁 Archivo Excel: {EXCEL_FILE}")
    print("📷 Imágenes almacenadas en base64 dentro del Excel")
    print("🌐 Abre tu navegador en: http://127.0.0.1:5000")
    print("=" * 50)
    app.run(debug=True, host='0.0.0.0', port=5000)
